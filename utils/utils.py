from __future__ import print_function
import torch
from PIL import Image
import inspect, re
import numpy as np
import os
import collections
import json
import csv
from skimage.exposure import rescale_intensity
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import pandas as pd
from shutil import rmtree

# Converts a Tensor into a Numpy array
# |imtype|: the desired type of the converted numpy array
def tensor2im(image_tensor, imgtype='img', datatype=np.uint8, batch_index=0):
    image_numpy = image_tensor[batch_index].cpu().float().numpy()
    if image_numpy.ndim == 4:# image_numpy (C x W x H x S)
        mid_slice = image_numpy.shape[-1]//2
        image_numpy = image_numpy[:,:,:,mid_slice]
    if image_numpy.shape[0] == 1:
        image_numpy = np.tile(image_numpy, (3, 1, 1))
    image_numpy = np.transpose(image_numpy, (1, 2, 0))
    if imgtype == 'img':
        image_numpy = (image_numpy + 8) / 16.0 * 255.0
    if np.unique(image_numpy).size == int(1):
        return image_numpy.astype(datatype)
    return rescale_intensity(image_numpy.astype(datatype))

def volume2img(volume):
    # Volume: np array
    # Todo add possibility to switch between mid_slice=True, labeled_slices=False
    def normalize(x):
        # clipped_x = np.clip(x, np.percentile(x, 1), np.percentile(x, 99)) # can be done fro signal enhancement
        clipped_x = x
        return np.subtract(clipped_x, np.min(clipped_x))/np.subtract(np.max(clipped_x), np.min(clipped_x))
    n_i, n_c, n_x, n_y, n_z = volume.shape
    center_z = n_z // 2
    for c in range(n_c):
        volume[:, c] = normalize(volume[:, c])

    return volume[:, :, :, :, center_z]

def diagnose_network(net, name='network'):
    mean = 0.0
    count = 0
    for param in net.parameters():
        if param.grad is not None:
            mean += torch.mean(torch.abs(param.grad.data))
            count += 1
    if count > 0:
        mean = mean / count
    print(name)
    print(mean)

def save_volumes(volumes, ids, save_dir, visualisation_format='png'):
    n_c = volumes['input'].shape[1]
    n_cols = 4
    for i in range(len(ids)):
        plt.figure()
        for c in range(n_c):
            plt.subplot(n_c, n_cols, c*n_cols + 1)
            plt.imshow(volumes['input'][i, c], cmap='gray')
            plt.title(f'channel {str(c)}')
            plt.axis('off')
        plt.subplot(n_c, n_cols, 2)
        plt.imshow(volumes['output'][i, 0], cmap='gray', vmin=0, vmax=1)
        plt.title('output')
        plt.axis('off')
        plt.subplot(n_c, n_cols, 3)
        plt.imshow(volumes['target'][i, 0], cmap='gray', vmin=0, vmax=1)
        plt.title('ground truth')
        plt.axis('off')
        plt.subplot(n_c, n_cols, 4)
        plt.imshow(volumes['target'][i, 0], cmap='gray', vmin=0, vmax=1)
        plt.imshow(volumes['output'][i, 0], cmap='Blues', alpha=0.6)
        plt.axis('off')
        plt.title('output + target')
        plt.suptitle(f'{ids[i]}')
        plt.savefig(save_dir + '/{}.{}'.format(ids[i], visualisation_format), format=visualisation_format)
        plt.close()


def save_image(image_numpy, image_path):
    image_pil = Image.fromarray(image_numpy)
    image_pil.save(image_path)


def info(object, spacing=10, collapse=1):
    """Print methods and doc strings.
    Takes module, class, list, dictionary, or string."""
    methodList = [e for e in dir(object) if isinstance(getattr(object, e), collections.Callable)]
    processFunc = collapse and (lambda s: " ".join(s.split())) or (lambda s: s)
    print( "\n".join(["%s %s" %
                     (method.ljust(spacing),
                      processFunc(str(getattr(object, method).__doc__)))
                     for method in methodList]) )

def varname(p):
    for line in inspect.getframeinfo(inspect.currentframe().f_back)[3]:
        m = re.search(r'\bvarname\s*\(\s*([A-Za-z_][A-Za-z0-9_]*)\s*\)', line)
        if m:
            return m.group(1)

def print_numpy(x, val=True, shp=False):
    x = x.astype(np.float64)
    if shp:
        print('shape,', x.shape)
    if val:
        x = x.flatten()
        print('mean = %3.3f, min = %3.3f, max = %3.3f, median = %3.3f, std=%3.3f' % (
            np.mean(x), np.min(x), np.max(x), np.median(x), np.std(x)))


def mkdirs(paths):
    if isinstance(paths, list) and not isinstance(paths, str):
        for path in paths:
            mkdir(path)
    else:
        mkdir(paths)


def mkdir(path):
    if not os.path.exists(path):
        os.makedirs(path)


def json_file_to_pyobj(filename):
    def _json_object_hook(d): return collections.namedtuple('X', d.keys())(*d.values())
    def json2obj(data): return json.loads(data, object_hook=_json_object_hook)
    return json2obj(open(filename).read())


def determine_crop_size(inp_shape, div_factor):
    div_factor= np.array(div_factor, dtype=np.float32)
    new_shape = np.ceil(np.divide(inp_shape, div_factor)) * div_factor
    pre_pad = np.round((new_shape - inp_shape) / 2.0).astype(np.int16)
    post_pad = ((new_shape - inp_shape) - pre_pad).astype(np.int16)
    return pre_pad, post_pad


def csv_write(out_filename, in_header_list, in_val_list):
    with open(out_filename, 'w') as f:
        writer = csv.writer(f)
        writer.writerow(in_header_list)
        writer.writerows(zip(*in_val_list))

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    # if startrow == 0:
        # df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
# R    else:
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

def save_config(json_opts, json_filename, model, val_loss_log, best_loss):
    """
    function that saves the experiment results and tests the model
    """

    # get the model path with the epoch of the best model
    model_path = os.path.join(model.save_dir, '{0:03d}_net_{1}.pth'.format(
        val_loss_log.loc[val_loss_log['Seg_Loss'] == best_loss, 'epoch'].item(),
        json_opts.model.model_type))

    # save config with path of trained model
    with open(json_filename) as file:
        config = json.load(file)
    config['model']['path_pre_trained_model'] = model_path
    config['model']['isTrain'] = False
    config_trained_path = os.path.join(model.save_dir, 'trained_' + json_filename.split('/')[-1])
    with open(config_trained_path, 'w') as outfile:
        json.dump(config, outfile, indent=4)

    return model_path

def rm_and_mkdir(path):
    if os.path.exists(path):
        print('removing dir ', path)
        rmtree(path)
    print('creating dir ', path)
    os.makedirs(path)